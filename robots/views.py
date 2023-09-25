import openpyxl

from django.db.models import Count
from django.utils import timezone
from django.http import JsonResponse, HttpResponse

from robots.models import Robot


def get_report(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'method not allowed'}, status=405)
    # Создем Excel файл и удаляем начальную страницу
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    start_date = timezone.now() - timezone.timedelta(days=7)

    result = (
        Robot.objects
        .filter(created__gte=start_date)
        .values("serial").values_list('model', 'version', named=True)
        .annotate(week_count=Count("model"))
    )
    # Получаем уникальные модели
    models = result.values_list('model', flat=True).distinct()

    for model in models:
        detail_data = result.filter(model=model)

        sheet = workbook.create_sheet(title=f"Модель {model}")
        sheet['A1'] = "Модель"
        sheet['B1'] = "Версия"
        sheet['C1'] = "Количество за неделю"

        row_num = 2
        for row in detail_data:
            sheet.cell(row=row_num, column=1, value=row[0])
            sheet.cell(row=row_num, column=2, value=row[1])
            sheet.cell(row=row_num, column=3, value=row[2])
            row_num += 1

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="robot_weekly_report.xlsx"'
    workbook.save(response)
    return response
